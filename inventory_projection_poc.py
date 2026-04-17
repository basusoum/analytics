import pandas as pd
import numpy as np
import os
import random
import shutil
import json
from datetime import date
from collections import defaultdict
from rca_rules_engine import compute_rca
from agentic_rca import generate_agentic_insights


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
_BASE        = os.path.dirname(os.path.abspath(__file__))
DATA_DIR     = os.path.join(_BASE, "generated_data")
DATA_FILE    = os.path.join(DATA_DIR, "synthetic_data.csv")
OUTPUT_FILE  = os.path.join(DATA_DIR, "final_projection_results__aligned.csv")
PIVOT_OUTPUT = os.path.join(DATA_DIR, "pivot_visualization.csv")
EXCEL_OUTPUT = os.path.join(DATA_DIR, "pulse_pivot.xlsx")
# Path to frontend public data directory for synchronization
FRONTEND_DATA_DIR = os.path.join(_BASE, "..", "frontend", "frontend", "public", "data")

TOTAL_PROJECTION_WEEKS = 12
PRIORITY_RANK          = {"P1": 1, "P2": 2, "P3": 3}
ALERT_BATCH_ID         = "24536581"

# Week label → real date mapping (Week_1 = 2026-02-08, weekly increments)
_WEEK_BASE = pd.Timestamp("2026-02-08")
WEEK_DATE  = {
    f"Week_{i}": (_WEEK_BASE + pd.Timedelta(weeks=i - 1)).strftime("%Y-%m-%d")
    for i in range(1, 13)
}
# Default Geographical Metadata for Sites
GEO_METADATA = {
    "FAC_01": {"Location": "Chicago",    "Region": "Midwest", "Lat": 41.8781, "Long": -87.6298},
    "DC_01":  {"Location": "Gary",       "Region": "Midwest", "Lat": 41.5934, "Long": -87.3464},
    "DC_02":  {"Location": "Aurora",     "Region": "Midwest", "Lat": 41.7606, "Long": -88.3201},
    "DC_03":  {"Location": "Naperville", "Region": "Midwest", "Lat": 41.7508, "Long": -88.1535},
    "DC_04":  {"Location": "Joliet",     "Region": "Midwest", "Lat": 41.5250, "Long": -88.0817},
    "DC_05":  {"Location": "Elgin",      "Region": "Midwest", "Lat": 42.0350, "Long": -88.2826},
}

# ── Excel pivot layout ────────────────────────────────────────────────────────
# Factory block: 15 figures (up to Excess only)
FIGURES_FACTORY = [
    ("Start Inventory",                    "Inv"),
    ("Total Forecast/Order",               "Demand"),
    ("Start Backorder",                    "Demand"),
    ("Total Demand",                       "Demand"),
    ("Planned Prod",                       "Supply"),
    ("Confirm Prod",                       "Supply"),
    ("Total Supply",                       "Supply"),
    ("End. Backorder",                     "Demand"),
    ("End. Inventory",                     "Inv"),
    ("MSTN Thresold",                      "Inv"),
    ("ESTN Thresold",                      "Inv"),
    ("MSTN%",                              ""),
    ("ESTN%",                              ""),
    ("Shortage",                           "Inv"),
    ("Excess",                             "Inv"),
    ("Out of stock Value($)",              "Val"),
    ("Excess and over($)",                 "Val"),
]

# DC block: 16 figures (up to Excess + Value impacts)
FIGURES_DC = [
    ("Start Inventory",                    "Inv"),
    ("Total Forecast/Order",               "Demand"),
    ("Start Backorder",                    "Demand"),
    ("Total Demand",                       "Demand"),
    ("Dist. Receipt Planned",              "Supply"),
    ("Total Supply",                       "Supply"),
    ("End. Backorder",                     "Demand"),
    ("End. Inventory",                     "Inv"),
    ("MSTN Thresold",                      "Inv"),
    ("ESTN Thresold",                      "Inv"),
    ("MSTN%",                              ""),
    ("ESTN%",                              ""),
    ("Shortage",                           "Inv"),
    ("Excess",                             "Inv"),
    ("Out of stock Value($)",              "Val"),
    ("Excess and over($)",                 "Val"),
]

FIG_IDX_FACTORY = {fig: i for i, (fig, _) in enumerate(FIGURES_FACTORY)}
FIG_IDX_DC      = {fig: i for i, (fig, _) in enumerate(FIGURES_DC)}
ROWS_FACTORY    = len(FIGURES_FACTORY) + 1   # 27 rows/block (26 + separator)
ROWS_DC         = len(FIGURES_DC)      + 1   # 23 rows/block (22 + separator)
FIXED_COLS     = ["SKU", "Site ID", "Site Type", "Figure Name", "Category"]
WEEK_COL_START = len(FIXED_COLS) + 1


# ══════════════════════════════════════════════════════════════════════════════
# FLAG / ALERT HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _mstn_flag_level(mstn_pct):
    """
     MSTN Deviation Flag Matrix:
      Critical : = 0%
      High     : > 0%  and <= 75%
      Medium   : > 75% and <  90%
      Low      : >= 90% and < 100%
      None     : = 100%  (no shortage)
    """
    if mstn_pct <= 0:   return "Critical"
    if mstn_pct <= 75:  return "High"
    if mstn_pct < 90:   return "Medium"
    if mstn_pct < 100:  return "Low"
    return "None"


def _estn_flag_level(estn_pct):
    """
     ESTN thresholds:
      High   : > 20%
      Medium : > 10% and <= 20%
      Low    : > 0%  and <= 10%
      None   : = 0%  (no excess)
    """
    if estn_pct > 20:  return "High"
    if estn_pct > 10:  return "Medium"
    if estn_pct > 0:   return "Low"
    return "None"


def _estn_priority(flag_level, persistence_pct):
    """
    ESTN Priority Matrix:

    | ESTN Flag    | >30–50% of weeks | >50% of weeks |
    |--------------|------------------|---------------|
    | Low/Moderate | P3               | P2            |
    | High         | P2               | P1            |

    Returns (alert_name, priority) or (None, None) if below 30% threshold.
    """
    if flag_level == "None" or persistence_pct <= 0.30:
        return None, None

    if persistence_pct > 0.50:
        if flag_level == "High":
            return "Critical Excess",   "P1"
        return "At-Risk Excess",        "P2"
    else:
        if flag_level == "High":
            return "At-Risk Excess",    "P2"
        return "Potential Excess",      "P3"


# ══════════════════════════════════════════════════════════════════════════════
# PROJECTION ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def _project_sku_pass1(sku, dc, sku_weekly, first):
    """Pass 1 — compute week-by-week inventory projections and per-week ESTN/MSTN flags."""
    ss          = first["Safety Stock (SS)"]
    ms          = first["Maximum Stock (Max)"]
    lead_time   = first["Replenishment / transportation lead time"]
    unit_price  = first.get("Unit Price", 0)
    
    # ≡ƒƒó Geo-Intelligence Fallback
    site_id_val = str(dc).strip()
    geo = GEO_METADATA.get(site_id_val, {})
    
    location    = first.get("Location") or geo.get("Location") or "N/A"
    region      = first.get("Region") or geo.get("Region") or "N/A"
    lat         = float(first.get("Latitude") if pd.notna(first.get("Latitude")) and first.get("Latitude") != 0 else geo.get("Lat", 0.0))
    lon         = float(first.get("Longitude") if pd.notna(first.get("Longitude")) and first.get("Longitude") != 0 else geo.get("Long", 0.0))
    
    site_hier   = first.get("Site", "DC")
    if isinstance(site_hier, str):
        site_hier = site_hier.strip().title() # Normalize to Title Case (Factory/Dc)
        if site_hier == "Dc": site_hier = "DC"
    fac_id      = first.get("Factory ID", "F1")
    dc_id_hier  = first.get("DC_IDS", "")
    comm_id     = first.get("Common id", "")
    initial_inv = first.get("Current on-hand inventory quantity", 0)
    current_inv = max(0, initial_inv)
    prev_end_backorder = max(0, -initial_inv)

    sku_rows           = []

    for _, week_row in sku_weekly.iterrows():
        opening_inv    = current_inv
        supply         = week_row["Confirmed supply quantities"]
        planned_supply = week_row.get("Planned supply quantities", 0)
        demand         = week_row["Forecast quantity"]

        start_backorder = prev_end_backorder
        total_demand    = demand + start_backorder
        total_supply    = supply + planned_supply
        
        net_position    = opening_inv + total_supply - total_demand
        closing_inv     = max(0, net_position)
        end_backorder   = max(0, -net_position)
        
        days_coverage   = round(closing_inv / (demand / 7), 1) if demand > 0 else 0

        prev_end_backorder = end_backorder

        mstn_pct = max(0, min(100, (net_position / ss) * 100)) if ss > 0 else 100
        estn_qty = max(0, net_position - ms)
        estn_pct = max(0, (estn_qty / ms) * 100) if ms > 0 else 0   # uncapped — BD allows >100%

        week_idx     = week_row["Week_Index"]
        is_within_lt = week_idx <= lead_time

        # MSTN alert —  priority matrix
        if mstn_pct <= 0:
            mstn_alert = ("Critical Stockout", "P1")
        elif mstn_pct <= 75:
            mstn_alert = ("Critical Stockout", "P1") if is_within_lt else ("At-Risk Stockout", "P2")
        elif mstn_pct < 100:
            mstn_alert = ("At-Risk Stockout", "P2") if is_within_lt else ("Potential Stockout", "P3")
        else:
            mstn_alert = None

        sku_rows.append({
            "sku": sku, "sku_desc": first["Material description"],
            "dc": dc, "bu": first["Business Unit"],
            "site_desc":  first.get("Site Description", dc),
            "site_type":  first.get("Location type",    "DC"),
            "category":   first.get("Category",         ""),
            "ss": ss, "ms": ms, "lead_time": lead_time,
            "ss_days":  first.get("Safety Stock (Days)", 28),
            "cs_units": first.get("Cycle Stock (Units)",  0),
            "cs_days":  first.get("Cycle Stock (Days)",   0),
            "ms_days":  first.get("Maximum Stock (Days)", 0),
            "week":             week_row["Week"],
            "week_idx":         week_idx,
            "opening_inv":      opening_inv,
            "supply":           supply,
            "planned_supply":   planned_supply,
            "total_supply":     total_supply,
            "demand":           demand,
            "closing_inv":      closing_inv,
            "start_backorder":  int(start_backorder),
            "end_backorder":    int(end_backorder),
            "total_demand":     int(total_demand),
            "days_coverage":    days_coverage,
            "mstn_pct":         int(mstn_pct),
            "estn_pct":         int(estn_pct),
            "estn_qty":         estn_qty,
            "shortage_units":   max(0, ss - net_position),
            "excess_units":     max(0, net_position - ms),
            "unit_price":       unit_price,
            "value_impact":     max(0, ss - net_position) * unit_price if (ss - net_position) > 0 else max(0, net_position - ms) * unit_price,
            "Location":         location,
            "Region":           region,
            "SKU_Category":     first.get("SKU Category", "C"),
            "Category":         str(first.get("Category", "Medical Devices")) if pd.notna(first.get("Category")) else "Medical Devices",
            "lat":              lat,
            "lon":              lon,
            "site":             site_hier,
            "factory_id":       fac_id,
            "dc_ids":           dc_id_hier,
            "common_id":        comm_id,
            "sku_category":     first.get("SKU Category", "C"),
            "sku_importance":   int(first.get("SKU Importance", 1)),
            "mstn_flag":        _mstn_flag_level(mstn_pct),
            "estn_flag":        _estn_flag_level(estn_pct),
            "mstn_alert":       mstn_alert,
            "is_within_lt":     is_within_lt,
        })

        current_inv = closing_inv

    return sku_rows


def _project_sku_pass2(sku_rows):
    """Pass 2 — ESTN persistence across the 12-week horizon."""
    total_weeks  = len(sku_rows)
    high_weeks   = sum(1 for r in sku_rows if r["estn_flag"] == "High")
    mod_or_above = sum(1 for r in sku_rows if r["estn_flag"] in ("Medium", "High"))
    any_estn     = sum(1 for r in sku_rows if r["estn_flag"] != "None")

    for r in sku_rows:
        flag = r["estn_flag"]
        if flag == "High":
            persistence_pct = high_weeks / total_weeks
        elif flag == "Moderate":
            persistence_pct = mod_or_above / total_weeks
        elif flag == "Low":
            persistence_pct = any_estn / total_weeks
        else:
            persistence_pct = 0.0

        estn_name, estn_pri = _estn_priority(flag, persistence_pct)
        r["estn_alert"]      = (estn_name, estn_pri) if estn_name else None
        r["persistence_pct"] = persistence_pct


def _project_sku_pass3(sku, dc, sku_rows):
    """Pass 3 — combine MSTN + ESTN alerts, pick highest priority, run RCA."""
    avg_demand      = sum(r["demand"] for r in sku_rows) / len(sku_rows) if sku_rows else 0
    shortage_streak = 0
    excess_streak   = 0

    for r in sku_rows:
        triggered = [a for a in (r["mstn_alert"], r["estn_alert"]) if a]

        if triggered:
            alert_status, priority = min(triggered, key=lambda x: PRIORITY_RANK[x[1]])
        else:
            alert_status, priority = "Healthy", "None"

        r["alert"]    = alert_status
        r["priority"] = priority

        is_shortage = "Stockout" in alert_status or "Shortage" in alert_status
        is_excess   = "Excess"   in alert_status

        if is_shortage:
            shortage_streak += 1; excess_streak = 0
        elif is_excess:
            excess_streak += 1; shortage_streak = 0
        else:
            shortage_streak = 0; excess_streak = 0

        consecutive = shortage_streak if is_shortage else excess_streak if is_excess else 0

        pipeline = [
            {"supply": x["supply"], "demand": x["demand"], "label": x["week"]}
            for x in sku_rows if x["week_idx"] > r["week_idx"]
        ]

        next_r     = next((x for x in sku_rows if x["week_idx"] == r["week_idx"] + 1), None)
        nxt_supply = next_r["supply"] if next_r else None
        nxt_label  = next_r["week"]   if next_r else None

        if r["priority"] != "None":
            r["rca_summary"] = compute_rca(
                alert      = r["alert"],
                opening    = r["opening_inv"],
                supply     = r["total_supply"],
                demand     = r["total_demand"],
                closing    = r["closing_inv"],
                ss         = r["ss"],
                ms         = r["ms"],
                lt         = r["lead_time"],
                week_idx   = r["week_idx"],
                mstn_pct   = r["mstn_pct"],
                estn_pct   = r["estn_pct"],
                avg_demand = avg_demand,
            )
            
            # 🔹 Agentic RCA Integration (Full)
            print(f"Generating Agentic RCA for {sku} at {dc} (Week {r['week_idx']})...")
            r["agentic_rca"] = generate_agentic_insights(sku, dc, r, {})
            # r["agentic_rca"] = ""
        else:
            r["rca_summary"] = ""
            r["agentic_rca"] = ""

        # Separate MSTN / ESTN comments for Alert Table (Sheet 4)
        r["mstn_rca"] = compute_rca(
            alert="Critical Stockout", opening=r["opening_inv"], supply=r["total_supply"],
            demand=r["total_demand"], closing=r["closing_inv"], ss=r["ss"], ms=r["ms"],
            lt=r["lead_time"], week_idx=r["week_idx"], avg_demand=avg_demand,
        ) if r["mstn_alert"] else ""

        r["estn_rca"] = compute_rca(
            alert="Critical Excess", opening=r["opening_inv"], supply=r["total_supply"],
            demand=r["total_demand"], closing=r["closing_inv"], ss=r["ss"], ms=r["ms"],
            lt=r["lead_time"], week_idx=r["week_idx"], avg_demand=avg_demand,
        ) if r["estn_alert"] else ""


def _to_final_row(r, snapshot_date, alert_id):
    """Flatten a projection row dict into the final CSV-ready dict."""
    alert_type = (
        f"{r['mstn_flag']} MSTN" if "Stockout" in r["alert"] or "Shortage" in r["alert"]
        else f"{r['estn_flag']} ESTN" if "Excess" in r["alert"]
        else "None"
    )

    # Priority Score (0–100 scale)
    # = 0.40*W1 + 0.30*W2 + 0.20*W3 + 0.10*W4
    # W1 – Priority Flag     : P1=100, P2=67, P3=33  (40% weight)
    # W2 – Shortage/Excess % : INT(qty/ref×100)       (30% weight)
    # W3 – Days to Alert     : INT(100/week_idx)       (20% weight)
    # W4 – SKU Importance    : A=100, B=67, C=33       (10% weight)
    cat = r.get("sku_category", "C")
    imp = r.get("sku_importance", 1)
    if r["priority"] != "None":
        is_shortage = "Stockout" in r["alert"] or "Shortage" in r["alert"]
        qty       = int(max(0, r["shortage_units"])) if is_shortage else int(max(0, r["excess_units"]))
        ref       = r["ss"] if is_shortage else r["ms"]
        w1_flag   = {"P1": 100, "P2": 67, "P3": 33}.get(r["priority"], 0)
        w2_qty    = int(min(100, (qty / ref) * 100)) if ref > 0 else 0
        w3_time   = int((1 / r["week_idx"]) * 100)
        w4_imp    = int((imp / 3) * 100)
        priority_score = int(round(0.40 * w1_flag + 0.30 * w2_qty + 0.20 * w3_time + 0.10 * w4_imp))
    else:
        w1_flag = w2_qty = w3_time = w4_imp = priority_score = 0

    return {
        "Alert_ID":                          alert_id,
        "Snapshot_Date":                     snapshot_date,
        "Week":                              WEEK_DATE.get(r["week"], r["week"]),
        "Week_Index":                        r["week_idx"],
        "SKU":                               r["sku"],
        "Site ID":                           r["dc"],
        "Site":                              r["dc"],
        "Site Type":                         r["site"],
        "Factory ID":                        r["factory_id"],
        "DC_IDS":                            r["dc_ids"],
        "Common id":                         r["common_id"],
        "Location":                          r.get("Location", r.get("location", "N/A")),
        "Region":                            r.get("Region",   r.get("region",   "N/A")),
        "Latitude":                          r["lat"],
        "Longitude":                         r["lon"],
        "Category":                          r["category"],
        "Unit Price":                        r.get("unit_price", 0),
        "Out of stock Value($)":             int(max(0, r["shortage_units"]) * r.get("unit_price", 0)),
        "Excess and over($)":                int(max(0, r["excess_units"]) * r.get("unit_price", 0)),
        "Start Inventory":                   int(r["opening_inv"]),
        "Total Forecast/Order":              int(r["demand"]),
        "Start Backorder":                   r["start_backorder"],
        "Total Demand":                      r["total_demand"],
        "Planned Prod":                      int(r["planned_supply"]),
        "Confirm Prod":                      int(r["supply"]),
        "Total Supply":                      int(r["total_supply"]),
        "End. Backorder":                    r["end_backorder"],
        "End. Inventory":                    int(r["closing_inv"]),
        "Days of Coverage":                  r["days_coverage"],
        "Lead Time":                         int(r["lead_time"]),
        "Target Safety Stock (in Days)":     int(r["ss_days"]),
        "Target Safety Stock (in Units)":    int(r["ss"]),
        "Cycle Stock (in Days)":             int(r["cs_days"]),
        "Cycle Stock (in Units)":            int(r["cs_units"]),
        "Target Max Stock (in Days)":        int(r["ms_days"]),
        "Target Max Stock (in Units)":       int(r["ms"]),
        "MSTN Thresold":                     int(r["ss"]),
        "ESTN Thresold":                     int(r["ms"]),
        "MSTN%":                             r["mstn_pct"],
        "ESTN%":                             r["estn_pct"],
        "Shortage":                          int(max(0, r["shortage_units"])),
        "Excess":                            int(max(0, r["excess_units"])),
        "Business Unit":                     r["bu"],
        "Alert Type":                        alert_type,
        "Priority":                          r["priority"],
        "Alert":                             r["alert"],
        "SKU_Category":                      cat,
        "W1_Priority_Flag":                  w1_flag,
        "W2_Shortage_Excess_Qty":            w2_qty,
        "W3_Days_to_Alert":                  w3_time,
        "W4_SKU_Importance":                 w4_imp,
        "Priority_Score":                    priority_score,
        "Agentic_RCA":                       r.get("agentic_rca", ""),
    }


# ══════════════════════════════════════════════════════════════════════════════
# PIVOT BUILDER (CSV)
# ══════════════════════════════════════════════════════════════════════════════

FIGURE_MAP = [
    ("Start Inventory",                "Inv",    lambda r: int(r["opening_inv"])),
    ("Total Forecast/Order",           "Demand", lambda r: int(r["demand"])),
    ("Start Backorder",                "Demand", lambda r: int(r["start_backorder"])),
    ("Total Demand",                   "Demand", lambda r: int(r["total_demand"])),
    ("Dist. Receipt Planned",          "Supply", lambda r: int(r["supply"])),
    ("Total Supply",                   "Supply", lambda r: int(r["total_supply"])),
    ("End. Backorder",                 "Demand", lambda r: int(r["end_backorder"])),
    ("End. Inventory",                 "Inv",    lambda r: int(r["closing_inv"])),
    ("Days of Coverage",               "Inv",    lambda r: r["days_coverage"]),
    ("Target Safety Stock (in Days)",  "Inv",    lambda r: int(r["ss_days"])),
    ("Target Safety Stock (in Units)", "Inv",    lambda r: int(r["ss"])),
    ("Cycle Stock (in Days)",          "Inv",    lambda r: int(r["cs_days"])),
    ("Cycle Stock (in Units)",         "Inv",    lambda r: int(r["cs_units"])),
    ("Target Max Stock (in Days)",     "Inv",    lambda r: int(r["ms_days"])),
    ("Target Max Stock (in Units)",    "Inv",    lambda r: int(r["ms"])),
    ("MSTN Thresold",                  "Inv",    lambda r: int(r["ss"])),
    ("ESTN Thresold",                  "Inv",    lambda r: int(r["ms"])),
    ("MSTN%",                          "",       lambda r: round(r["mstn_pct"] / 100, 6)),
    ("ESTN%",                          "",       lambda r: round(r["estn_qty"] / r["ms"], 6) if r["ms"] > 0 else 0),
    ("Shortage",                       "Inv",    lambda r: int(max(0, r["shortage_units"]))),
    ("Excess",                         "Inv",    lambda r: int(max(0, r["excess_units"]))),
    ("Out of stock Value($)",          "Val",    lambda r: int(max(0, r["shortage_units"]) * r.get("unit_price", 0))),
    ("Excess and over($)",             "Val",    lambda r: int(max(0, r["excess_units"]) * r.get("unit_price", 0))),
]


def _build_csv_pivot(all_results):
    sku_site_groups = defaultdict(dict)
    for r in all_results:
        wd = WEEK_DATE.get(r["week"], r["week"])
        sku_site_groups[(r["sku"], r["dc"])][wd] = r

    week_cols  = [WEEK_DATE[f"Week_{i}"] for i in range(1, TOTAL_PROJECTION_WEEKS + 1)]
    # Simplified base columns as requested: SKU, Site, Metadata, and Figure Name
    base_cols  = ["SKU", "Site ID", "Site", "Site Type", "Region", "Location", "Category", "SKU_Category", "Figure Name"]
    pivot_rows = []

    for (sku, dc), week_data in sku_site_groups.items():
        first_r = next(iter(week_data.values()))
        for fig_name, cat, val_fn in FIGURE_MAP:
            row = {
                "SKU":          sku,
                "Site ID":      dc,
                "Site":         dc,
                "Site Type":    first_r.get("Site Type", first_r.get("site", "DC")),
                "Region":       first_r.get("Region",   "Midwest"),
                "Location":     first_r.get("Location", "Chicago"),
                "Category":     first_r.get("Category", "Medical Devices"),
                "SKU_Category": first_r.get("SKU_Category", "C"),
                "Figure Name":  fig_name,
            }
            for wd in week_cols:
                r = week_data.get(wd)
                row[wd] = val_fn(r) if r else ""
            pivot_rows.append(row)
        pivot_rows.append(dict.fromkeys(base_cols + week_cols, ""))   # blank separator

    df_pivot = pd.DataFrame(pivot_rows, columns=base_cols + week_cols)
    df_pivot.to_csv(PIVOT_OUTPUT, index=False, encoding="utf-8-sig")
    print(f"\nVisualization pivot saved to: {PIVOT_OUTPUT}")
    print(f"Pivot shape: {df_pivot.shape} ({len(sku_site_groups)} SKU-Site blocks × {len(FIGURE_MAP)} rows each)")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_excel_pivot(all_results, final_rows, alert_table_rows, output_path):
    """
    Generates pulse_pivot.xlsx with 2 sheets:

    Sheet 1 "POC-Network Overview":
      - Data blocks grouped by SKU, Factory site first then DCs alphabetically
      - BD exact formulas replicated:
          End. Backorder : =-MIN(0, SI+TS-TD)
          End. Inventory : =MAX(SI+TS-TD, 0)
          MSTN%          : =IFERROR(MIN(EI,MT)/MT, 1)
          ESTN%          : =MAX(EI-ET, 0)/ET
      - 8 BD alert rows per block

    Sheet 2 "Transposed View":
      - Flat format: weeks as rows, all 23 metrics/alert values as columns
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    def _format_sheet(ws):
        """Bold header row + auto-fit all column widths."""
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=8
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "POC-Network Overview"

    week_dates  = [WEEK_DATE[f"Week_{i}"] for i in range(1, TOTAL_PROJECTION_WEEKS + 1)]
    first_wk_cl = get_column_letter(WEEK_COL_START)
    last_wk_cl  = get_column_letter(WEEK_COL_START + TOTAL_PROJECTION_WEEKS - 1)

    HEADER_ROW = 1
    DATA_START = 2

    for ci, h in enumerate(FIXED_COLS + week_dates, 1):
        ws.cell(row=HEADER_ROW, column=ci, value=h)

    # ── Group: SKU → [(dc, site_type, week_data), ...] ───────────────────────
    sku_site_data = defaultdict(dict)
    for r in all_results:
        wd = WEEK_DATE.get(r["week"], r["week"])
        sku_site_data[(r["sku"], r["dc"])][wd] = r

    sku_groups = defaultdict(list)
    for (sku, dc), week_data in sku_site_data.items():
        first_r   = next(iter(week_data.values()))
        site_type = first_r.get("site_type", "DC")
        sku_groups[sku].append((dc, site_type, week_data))

    def _site_order(x):
        return (0 if x[1] == "Factory" else 1, x[0])

    # ── Pre-calculate block start rows for factory→DC pull-demand formulas ───
    block_start_rows = {}
    _temp_row = DATA_START
    for _sku in sorted(sku_groups.keys()):
        for _dc, _st, _ in sorted(sku_groups[_sku], key=_site_order):
            block_start_rows[(_sku, _dc)] = _temp_row
            _temp_row += ROWS_FACTORY if _st == "Factory" else ROWS_DC

    # ── Sheet 1 ───────────────────────────────────────────────────────────────
    excel_row = DATA_START

    for sku in sorted(sku_groups.keys()):
        for dc, site_type, week_data in sorted(sku_groups[sku], key=_site_order):
            first_r   = next(iter(week_data.values()))
            block     = excel_row
            lead_time = int(first_r.get("lead_time", 4))

            # Select per-site figures, index, and block size
            is_factory      = (site_type == "Factory")
            figures         = FIGURES_FACTORY if is_factory else FIGURES_DC
            fi              = FIG_IDX_FACTORY if is_factory else FIG_IDX_DC
            rows_this_block = ROWS_FACTORY    if is_factory else ROWS_DC

            # Row numbers — shared figures
            si = block + fi["Start Inventory"]
            fc = block + fi["Total Forecast/Order"]
            sb = block + fi["Start Backorder"]
            td = block + fi["Total Demand"]
            ts = block + fi["Total Supply"]
            eb = block + fi["End. Backorder"]
            ei = block + fi["End. Inventory"]
            mt = block + fi["MSTN Thresold"]
            et = block + fi["ESTN Thresold"]
            mp = block + fi["MSTN%"]
            ep = block + fi["ESTN%"]
            sh = block + fi["Shortage"]
            ex = block + fi["Excess"]

            # Factory-only row numbers
            if is_factory:
                pp = block + fi["Planned Prod"]
                cp = block + fi["Confirm Prod"]
            else:
                cp = block + fi["Dist. Receipt Planned"]

            # For Factory: collect DC Dist. Receipt Planned row numbers for pull-demand formula
            dc_cp_rows = []
            if is_factory:
                for dc_k, st_k, _ in sorted(sku_groups[sku], key=_site_order):
                    if st_k != "Factory":
                        dc_cp_rows.append(
                            block_start_rows[(sku, dc_k)] + FIG_IDX_DC["Dist. Receipt Planned"]
                        )

            # Fixed left columns — loop over site-specific figure list (no blank rows)
            for fig_name, cat in figures:
                row      = block + fi[fig_name]
                is_first = (fig_name == "Start Inventory")
                ws.cell(row=row, column=1, value=sku if is_first else "")
                ws.cell(row=row, column=2, value=dc if is_first else "")
                ws.cell(row=row, column=3, value=site_type)
                ws.cell(row=row, column=4, value=fig_name)
                ws.cell(row=row, column=5, value=cat)

            # Week columns with BD exact formulas
            for wk_idx, wd in enumerate(week_dates):
                col     = WEEK_COL_START + wk_idx
                cl      = get_column_letter(col)
                prev_cl = get_column_letter(col - 1) if wk_idx > 0 else None
                r       = week_data.get(wd, {})

                ws.cell(row=si, column=col,
                        value=(int(first_r.get("opening_inv", 0)) if wk_idx == 0
                               else f"={prev_cl}{ei}"))

                # Factory: pull-demand formula = SUM(DC Dist. Receipt Planned for Week N+1)
                # DC: hardcoded forecast value
                if site_type == "Factory" and dc_cp_rows:
                    if wk_idx < TOTAL_PROJECTION_WEEKS - 1:
                        next_cl = get_column_letter(WEEK_COL_START + wk_idx + 1)
                        ws.cell(row=fc, column=col,
                                value="=" + "+".join(f"{next_cl}{dc_row}" for dc_row in dc_cp_rows))
                    else:
                        ws.cell(row=fc, column=col, value=0)   # no Week N+1 at horizon end
                else:
                    ws.cell(row=fc, column=col, value=int(r.get("demand", 0)))

                ws.cell(row=sb, column=col,
                        value=(0 if wk_idx == 0 else f"={prev_cl}{eb}"))

                ws.cell(row=td, column=col, value=f"={cl}{fc}+{cl}{sb}")
                if is_factory:
                    ws.cell(row=pp, column=col, value=int(r.get("planned_supply", 0)))
                    ws.cell(row=cp, column=col, value=int(r.get("supply", 0)))
                    ws.cell(row=ts, column=col, value=f"={cl}{pp}+{cl}{cp}")
                else:
                    ws.cell(row=cp, column=col, value=int(r.get("supply", 0)))
                    ws.cell(row=ts, column=col, value=f"={cl}{cp}")

                # BD formulas
                ws.cell(row=eb, column=col, value=f"=-MIN(0,{cl}{si}+{cl}{ts}-{cl}{td})")
                ws.cell(row=ei, column=col, value=f"=MAX({cl}{si}+{cl}{ts}-{cl}{td},0)")
                ws.cell(row=mt, column=col, value=int(first_r.get("ss", 0)))
                ws.cell(row=et, column=col, value=int(first_r.get("ms", 0)))

                c = ws.cell(row=mp, column=col,
                            value=f"=IFERROR(MIN({cl}{ei},{cl}{mt})/{cl}{mt},1)")
                c.number_format = "0%"

                c = ws.cell(row=ep, column=col,
                            value=f"=MAX({cl}{ei}-{cl}{et},0)/{cl}{et}")
                c.number_format = "0%"

                ws.cell(row=sh, column=col, value=f"=MAX(0,{cl}{mt}-{cl}{ei})")
                ws.cell(row=ex, column=col, value=f"=MAX(0,{cl}{ei}-{cl}{et})")

            excel_row += rows_this_block

    # ── Sheet 2: Transposed View ──────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Transposed View")

    SHEET2_COLS = [
        "Alert_ID", "Snapshot_Date", "Week", "Week_Index",
        "SKU", "Site ID", "Factory ID", "Location", "Region",
        "Latitude", "Longitude", "Site Type", "Category",
        "Start Inventory", "Total Forecast/Order", "Start Backorder", "Total Demand",
        "Planned Prod", "Confirm Prod", "Total Supply", "End. Backorder", "End. Inventory",
        "Days of Coverage", "Lead Time",
        "Target Safety Stock (in Days)", "Target Safety Stock (in Units)",
        "Cycle Stock (in Days)", "Cycle Stock (in Units)",
        "Target Max Stock (in Days)", "Target Max Stock (in Units)",
        "MSTN Thresold", "ESTN Thresold", "MSTN%", "ESTN%",
        "Shortage", "Excess",
        "Business Unit", "Alert Type", "Priority", "Alert",
        "SKU_Category", "SKU_Importance",
        "W1_Priority_Flag", "W2_Shortage_Excess_Qty", "W3_Days_to_Alert",
        "W4_SKU_Importance", "Priority_Score", "Root Cause Analysis/AI Insights",
    ]

    for ci, h in enumerate(SHEET2_COLS, 1):
        ws2.cell(row=1, column=ci, value=h)

    # Pre-compute column letters for formula columns
    s2_col   = {col: i for i, col in enumerate(SHEET2_COLS, 1)}
    pri_cl   = get_column_letter(s2_col["Priority"])
    wk_cl    = get_column_letter(s2_col["Week_Index"])
    ss_cl    = get_column_letter(s2_col["Target Safety Stock (in Units)"])
    ms_cl    = get_column_letter(s2_col["Target Max Stock (in Units)"])
    sht_cl   = get_column_letter(s2_col["Shortage"])
    exc_cl   = get_column_letter(s2_col["Excess"])
    cat_cl   = get_column_letter(s2_col["SKU_Category"])
    w1_cl    = get_column_letter(s2_col["W1_Priority_Flag"])
    w2_cl    = get_column_letter(s2_col["W2_Shortage_Excess_Qty"])
    w3_cl    = get_column_letter(s2_col["W3_Days_to_Alert"])
    w4_cl    = get_column_letter(s2_col["W4_SKU_Importance"])

    FORMULA_COLS = {"W1_Priority_Flag", "W2_Shortage_Excess_Qty",
                    "W3_Days_to_Alert", "W4_SKU_Importance", "Priority_Score"}

    for row2_idx, row_dict in enumerate(final_rows, 2):
        r = row2_idx
        for ci, col_name in enumerate(SHEET2_COLS, 1):
            if col_name not in FORMULA_COLS:
                # Use standard row mapping — handling a few common spelling diffs
                key = (
                    "sku" if col_name == "SKU"
                    else "dc"  if col_name == "Site"
                    else "lat" if col_name == "Latitude"
                    else "lon" if col_name == "Longitude"
                    else "Agentic_RCA" if col_name == "Root Cause Analysis/AI Insights"
                    else col_name.lower().replace(" ", "_")
                )
                ws2.cell(row=r, column=ci, value=row_dict.get(col_name, row_dict.get(key, "")))
                continue

            # W1 — Priority Flag: P1=100, P2=67, P3=33, else 0
            if col_name == "W1_Priority_Flag":
                ws2.cell(row=r, column=ci, value=(
                    f'=IF({pri_cl}{r}="P1",100,IF({pri_cl}{r}="P2",67,IF({pri_cl}{r}="P3",33,0)))'
                ))

            # W2 — Shortage or Excess as % of SS/Max
            elif col_name == "W2_Shortage_Excess_Qty":
                ws2.cell(row=r, column=ci, value=(
                    f'=MIN(100,INT(IF({sht_cl}{r}>0,'
                    f'{sht_cl}{r}/{ss_cl}{r},{exc_cl}{r}/{ms_cl}{r})*100))'
                ))

            # W3 — Week urgency: INT(100/Week_Index)
            elif col_name == "W3_Days_to_Alert":
                ws2.cell(row=r, column=ci, value=(
                    f'=IF({sht_cl}{r}+{exc_cl}{r}=0,0,INT(100/{wk_cl}{r}))'
                ))

            # W4 — SKU importance: A=100, B=67, C=33
            elif col_name == "W4_SKU_Importance":
                ws2.cell(row=r, column=ci, value=(
                    f'=IF({cat_cl}{r}="A",100,IF({cat_cl}{r}="B",67,33))'
                ))

            # Priority Score — 0.40*W1 + 0.30*W2 + 0.20*W3 + 0.10*W4
            elif col_name == "Priority_Score":
                ws2.cell(row=r, column=ci, value=(
                    f'=ROUND(0.40*{w1_cl}{r}+0.30*{w2_cl}{r}+0.20*{w3_cl}{r}+0.10*{w4_cl}{r},0)'
                ))

    # ── Sheet 3: Planner Triage View ──────────────────────────────────────────
    ws3 = wb.create_sheet(title="Planner Triage View")
    
    SHEET3_COLS = [
        "Priority", "SKU_Category", "Week", "Shortage/Excess %",
        "Alert Type", "SKU", "Site ID", "Root Cause Analysis/AI Insights"
    ]
    
    for ci, h in enumerate(SHEET3_COLS, 1):
        ws3.cell(row=1, column=ci, value=h)
        
    # Sort alerts specifically for Sheet 3: Priority (P1→P2→P3) → Category (A→B→C) → Gap% desc → Week asc
    _PRI_SORT = {"P1": 1, "P2": 2, "P3": 3}
    _CAT_SORT = {"A": 1, "B": 2, "C": 3}
    sheet3_rows = [r for r in final_rows if r.get("Priority") != "None"]
    sheet3_rows.sort(key=lambda r: (
        _PRI_SORT.get(r.get("Priority"), 9),
        _CAT_SORT.get(r.get("SKU_Category"), 9),
        -r.get("W2_Shortage_Excess_Qty", 0),
        r.get("Week_Index", 99),
    ))

    row3_idx = 2
    for row_dict in sheet3_rows:
        ws3.cell(row=row3_idx, column=1, value=row_dict.get("Priority", ""))
        ws3.cell(row=row3_idx, column=2, value=row_dict.get("SKU_Category", ""))
        ws3.cell(row=row3_idx, column=3, value=row_dict.get("Week_Index", ""))
        
        # Format the gap as a percentage string (e.g., "90%")
        gap_val = row_dict.get("W2_Shortage_Excess_Qty", 0)
        ws3.cell(row=row3_idx, column=4, value=f"{gap_val}%")
        
        ws3.cell(row=row3_idx, column=5, value=row_dict.get("Alert Type", ""))
        ws3.cell(row=row3_idx, column=6, value=row_dict.get("SKU", ""))
        ws3.cell(row=row3_idx, column=7, value=row_dict.get("Site", ""))
        ws3.cell(row=row3_idx, column=8, value=row_dict.get("Agentic_RCA", ""))
        
        row3_idx += 1

    # ── Sheet 4: Alert Table ──────────────────────────────────────────────────
    ws4 = wb.create_sheet(title="Alert Table")

    SHEET4_COLS = [
        "Snapshot_Date", "Alert_ID", "SKU", "Location", "Week_No",
        "MSTN%", "Shortage", "MST_Category", "MSTN_Priority", "MSTN_Comment",
        "ESTN%", "Excess", "ESTN_Category", "ESTN_Priority", "ESTN_Comment",
        "Root Cause Analysis/AI Insights",
    ]

    for ci, h in enumerate(SHEET4_COLS, 1):
        ws4.cell(row=1, column=ci, value=h)

    for row_idx, row_dict in enumerate(alert_table_rows, 2):
        for ci, col_name in enumerate(SHEET4_COLS, 1):
            key = "Agentic_RCA" if col_name == "Root Cause Analysis/AI Insights" else col_name
            ws4.cell(row=row_idx, column=ci, value=row_dict.get(key, ""))

    # Apply bold headers + auto-fit to all sheets
    for _ws in [ws, ws2, ws3, ws4]:
        _format_sheet(_ws)

    wb.save(output_path)
    print(f"\nExcel pivot saved to: {output_path}")
    print(f"  Sheet 1 'POC-Network Overview': {excel_row - DATA_START} data rows "
          f"| Factory blocks: {ROWS_FACTORY} rows ({len(FIGURES_FACTORY)} metrics + 1 sep) "
          f"| DC blocks: {ROWS_DC} rows ({len(FIGURES_DC)} metrics + 1 sep)")
    print(f"  Sheet 2 'Transposed View': {len(final_rows)} rows x {len(SHEET2_COLS)} columns")
    print(f"  Sheet 3 'Planner Triage View': {row3_idx - 2} active alerts")
    print(f"  Sheet 4 'Alert Table': {len(alert_table_rows)} rows x {len(SHEET4_COLS)} columns")


# ══════════════════════════════════════════════════════════════════════════════
# DATA STANDARDIZATION
# ══════════════════════════════════════════════════════════════════════════════

def standardize_columns(df):
    """
    Intelligently maps user-provided columns to the canonical names 
    expected by the projection engine.
    """
    mapping = {
        'Material / SKU ID': ['Material / SKU ID', 'SKU', 'Material', 'Item', 'Material ID', 'Material / SKU ID'],
        'Site ID': ['Site ID', 'Site', 'Location ID', 'DC ID', 'Factory ID', 'Site ID'],
        'Safety Stock (SS)': ['Safety Stock (SS)', 'Safety Stock', 'SS', 'Minimum Stock', 'Target Safety Stock (in Units)', 'Safety Stock (SS)'],
        'Maximum Stock (Max)': ['Maximum Stock (Max)', 'Maximum Stock', 'Max Stock', 'MS', 'Target Max', 'Target Max Stock (in Units)', 'Maximum Stock (Max)'],
        'Replenishment / transportation lead time': ['Replenishment / transportation lead time', 'Lead Time', 'LT', 'Lead_Time', 'LeadTime', 'Replenishment / transportation lead time'],
        'Forecast quantity': ['Forecast quantity', 'Forecast', 'Demand', 'Total Forecast/Order', 'Quantity', 'Forecast quantity'],
        'Current on-hand inventory quantity': ['Current on-hand inventory quantity', 'On-Hand', 'Stock', 'Inventory', 'Current Inventory', 'Current on-hand inventory quantity'],
        'Confirmed supply quantities': ['Confirmed supply quantities', 'Supply', 'Confirmed Supply', 'Receipts', 'Confirm Prod', 'Confirmed Prod', 'Confirmed supply quantities'],
        'Planned supply quantities': ['Planned supply quantities', 'Planned Supply', 'Projected Supply', 'Planned Prod', 'Planned Production', 'Planned supply quantities'],
        'Week_Index': ['Week_Index', 'Week Index', 'Week No', 'Week Number', 'Week_Index'],
        'Material description': ['Material description', 'Description', 'Material Name', 'Material description'],
        'Unit Price': ['Unit Price', 'Price', 'Price/Unit', 'Unit Price'],
        'Location': ['Location', 'City', 'Place', 'Location'],
        'Region': ['Region', 'Area', 'Zone', 'Region'],
        'Latitude': ['Latitude', 'Lat', 'Latitude'],
        'Longitude': ['Longitude', 'Long', 'Lon', 'Longitude'],
        'Site': ['Site', 'Site Type', 'Location Type', 'Site'],
        'Material / SKU ID': ['Material / SKU ID', 'SKU', 'Material', 'Item', 'Material ID']
    }

    cols = df.columns.tolist()
    rename_dict = {}

    for canonical, variations in mapping.items():
        # Skip if canonical is already in columns
        if canonical in cols:
            continue
            
        # Try to find a match among variations
        for var in variations:
            match = next((c for c in cols if c.strip().lower() == var.lower()), None)
            if match:
                rename_dict[match] = canonical
                break
    
    if rename_dict:
        print(f"Standardizing column names: {rename_dict}")
        return df.rename(columns=rename_dict)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def run_mass_projection(df=None):
    if df is None:
        if not os.path.exists(DATA_FILE):
            print("Error: synthetic_data.csv not found. Run generate_synthetic_data.py first.")
            return
        print("Loading synthetic data from disk...")
        df = pd.read_csv(DATA_FILE)
    else:
        print("Processing provided DataFrame...")
        # Standardize columns before saving and processing
        df = standardize_columns(df)
        
        # Save a copy as synthetic_data.csv for record keeping
        os.makedirs(DATA_DIR, exist_ok=True)
        df.to_csv(DATA_FILE, index=False)
    
    # Final check for critical columns after standardization
    critical_cols = ['Material / SKU ID', 'Site ID', 'Safety Stock (SS)', 'Forecast quantity']
    missing = [c for c in critical_cols if c not in df.columns]
    if missing:
        raise KeyError(f"Missing required columns (even after mapping): {missing}. Please check your data format.")

    # Assign SKU Category (A/B/C) — one per SKU, fixed with seed
    random.seed(42)
    all_skus       = sorted(df["Material / SKU ID"].unique())
    sku_cat_map    = {sku: random.choice(["A", "B", "C"]) for sku in all_skus}
    print(f"SKU Categories: {sku_cat_map}")

    all_results   = []
    snapshot_date = date.today().strftime("%Y-%m-%d")
    sku_dc_groups = df.groupby(["Material / SKU ID", "Site ID"])
    print(f"Processing {len(sku_dc_groups)} SKU-DC combinations... (Snapshot: {snapshot_date})")

    for (sku, dc), sku_weekly_df in sku_dc_groups:
        sku_weekly = sku_weekly_df.sort_values("Week_Index")
        first      = sku_weekly.iloc[0]

        sku_rows = _project_sku_pass1(sku, dc, sku_weekly, first)
        for row in sku_rows:
            row["sku_category"] = sku_cat_map.get(sku, "C")
        _project_sku_pass2(sku_rows)
        _project_sku_pass3(sku, dc, sku_rows)
        all_results.extend(sku_rows)

    # Assign Alert IDs and flatten to final rows
    alert_counter = 1
    final_rows    = []
    for r in all_results:
        if r["priority"] != "None":
            alert_id = f"BD-{ALERT_BATCH_ID}-{alert_counter:05d}"
            alert_counter += 1
        else:
            alert_id = ""
        final_rows.append(_to_final_row(r, snapshot_date, alert_id))

    # Save flat output
    df_results = pd.DataFrame(final_rows)
    df_results.to_csv(OUTPUT_FILE, index=False, encoding="utf-8-sig")
    print(f"\nSUCCESS: Projection complete.")
    print(f"Results saved to : {OUTPUT_FILE}")
    print(f"Total columns    : {len(df_results.columns)}")
    print("\nAlert Summary:")
    print(df_results["Alert"].value_counts())
    print("\nPriority Summary:")
    print(df_results["Priority"].value_counts())

    # Build Alert ID lookup for Sheet 4
    alert_id_lookup = {
        (fr["SKU"], fr["Site ID"], fr["Week_Index"]): fr.get("Alert_ID", "")
        for fr in final_rows
    }

    # Build Sheet 4 alert table rows — one row per SKU-Site-Week with MSTN or ESTN alert
    alert_table_rows = []
    for r in all_results:
        if not r["mstn_alert"] and not r["estn_alert"]:
            continue
        alert_table_rows.append({
            "Snapshot_Date":  snapshot_date,
            "Alert_ID":       alert_id_lookup.get((r["sku"], r["dc"], r["week_idx"]), ""),
            "SKU":            r["sku"],
            "Location":       r["dc"],
            "Week_No":        r["week_idx"],
            "MSTN%":          r["mstn_pct"] if r["mstn_alert"] else "",
            "Shortage":       int(max(0, r["shortage_units"])) if r["mstn_alert"] else "",
            "MST_Category":   r["mstn_flag"] if r["mstn_alert"] else "",
            "MSTN_Priority":  r["mstn_alert"][1] if r["mstn_alert"] else "",
            "MSTN_Comment":   r.get("mstn_rca", ""),
            "ESTN%":          r["estn_pct"] if r["estn_alert"] else "",
            "Excess":         int(max(0, r["excess_units"])) if r["estn_alert"] else "",
            "ESTN_Category":  r["estn_flag"] if r["estn_alert"] else "",
            "ESTN_Priority":  r["estn_alert"][1] if r["estn_alert"] else "",
            "ESTN_Comment":   r.get("estn_rca", ""),
            "Agentic_RCA":    r.get("agentic_rca", ""),
        })

    _build_csv_pivot(all_results)
    build_excel_pivot(all_results, final_rows, alert_table_rows, EXCEL_OUTPUT)

    # Save Alert Center Report specifically for easy access
    alert_report_xlsx = os.path.join(DATA_DIR, "Alert_Center_Report.xlsx")
    alert_report_csv  = os.path.join(DATA_DIR, "Alert_Center_Report.csv")
    pd.DataFrame(alert_table_rows).to_excel(alert_report_xlsx, index=False, engine='openpyxl')
    pd.DataFrame(alert_table_rows).to_csv(alert_report_csv, index=False)
    print(f"Alert report saved to: {alert_report_xlsx}")


def sync_to_frontend():
    """Copies generated data files to the frontend/public/data directory."""
    if not os.path.exists(FRONTEND_DATA_DIR):
        print(f"Warning: Frontend data directory not found at {FRONTEND_DATA_DIR}")
        return

    # Files to sync (mapping source name to target name as expected by useData hook)
    sync_map = {
        "pivot_visualization.csv": "pivot_visualization.csv",
        "final_projection_results__aligned.csv": "projection_results.csv",
        "Alert_Center_Report.csv": "Alert_Center_Report.csv",
        "pulse_pivot.xlsx": "pulse_pivot.xlsx"
    }

    print(f"Syncing data to frontend: {FRONTEND_DATA_DIR}")
    for src_name, dest_name in sync_map.items():
        src_path = os.path.join(DATA_DIR, src_name)
        dest_path = os.path.join(FRONTEND_DATA_DIR, dest_name)
        if os.path.exists(src_path):
            shutil.copy2(src_path, dest_path)
            print(f"  [OK] {src_name} -> {dest_name}")

    # Generate or update manifest.json in the frontend
    manifest_path = os.path.join(FRONTEND_DATA_DIR, "manifest.json")
    manifest = {
        "generatedAt": date.today().strftime("%Y-%m-%d %H:%M:%S"),
        "roles": {
            "pivot": "/data/pivot_visualization.csv",
            "projection": "/data/projection_results.csv",
            "triage": "/data/Alert_Center_Report.csv"
        }
    }
    with open(manifest_path, 'w') as f:
        json.dump(manifest, f, indent=2)
    print(f"  [OK] manifest.json updated")


if __name__ == "__main__":
    run_mass_projection()
    sync_to_frontend()